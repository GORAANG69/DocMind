"""XLSX parser using openpyxl."""
from pathlib import Path

from openpyxl import load_workbook

from parsers.base_parser import BaseParser


class XlsxParser(BaseParser):
    """Extract text from Excel XLSX files."""

    def extract_text(self, file_path: Path) -> str:
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX '{file_path.name}': {exc}") from exc

        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "":
                        parts.append(f"{sheet_name}\t{cell.coordinate}\t{str(cell.value).strip()}")

        wb.close()

        if not parts:
            raise ValueError(f"XLSX '{file_path.name}' contains no extractable data.")

        return "\n".join(parts)

    @staticmethod
    def supported_extensions() -> list[str]:
        return [".xlsx"]
